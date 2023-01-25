from faker import Faker
from openpyxl import Workbook
from openpyxl.styles import *

fak = Faker("ko-KR")

wb = Workbook()
ws = wb.active

 
ws.append(['이름','주소','회사','이메일','나이'])

for i in range(20):
    ws.append([fak.name(),fak.address(),fak.company(),fak.email(),fak.pyint(min_value=0, max_value=100)])

ws.column_dimensions['A'].width=20
ws.column_dimensions['B'].width=35
ws.column_dimensions['C'].width=20
ws.column_dimensions['D'].width=25

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"))

for cell in ws[1]:
    cell.font = Font(color="FF0000",size=20,italic=True,bold=True)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center",vertical="center")

# ws.freeze_panes = "B1"

ws.merge_cells("A1:D1")

from openpyxl.drawing.image import Image

img = Image('aa.png')
ws.add_image(img,'A20')

wb.save('randomperson.xlsx')
wb.close()