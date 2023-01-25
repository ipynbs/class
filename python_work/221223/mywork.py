from openpyxl import Workbook,load_workbook

class MyWorkBook:
    
    #생성자 MyWorkBook()
    def __init__(self):
        self.cnt = 1
        print("생성자 호출했음"+str(self.cnt))
        print(f"생성자 호출했음 {self.cnt}")

    def setCnt(self,cnt):
        self.cnt = cnt
        print("cnt바꿈")
    
    def doSave(self):
        wb = Workbook()
        ws = wb.active
        # 저장
        for r in range(1,self.cnt):
            for c in range(1,r+1):
                ws.cell(row = r, column=c,value='*')

        wb.save("star.xlsx")
        wb.close()

    def doLoad(self):
        #불러오기
        lb = load_workbook('star.xlsx')
        ws = lb.active
        for r in range(1,self.cnt):
            for c in range(1,r+1):
                print(ws.cell(row = r, column=c).value,end=" ")
            print()

        lb.close()