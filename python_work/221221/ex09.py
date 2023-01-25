from openpyxl import Workbook,load_workbook
import random

# 1. Workbook 파일 새로 생성
# 2. load_workbook 파일 열기
# 3. random 사용하면 랜덤 값 가져오기

'''
    문제
    star.xlsx 파일 저장하기 
    *
    * *
    * * *
    * * * *
    * * * * *
    star.xlsx 파일 불러오기
    *
    * *
    * * *
    * * * *
    * * * * *
    클래스명 Star 로 작성하기
    1. 메서드 makeStar()
    2. 메서드 loadStar()

    
'''
wb = Workbook()
ws = wb.active
#A10 ~ J10 
# 1행1열 10행10열까지 랜덤한 값을 넣어서
for x in range(1,11):
    for y in range(1,11):
        ws.cell(row=y,column=x).value = random.randint(0,100)
wb.save("random.xlsx")
wb.close()

#불러와서 출력하기
lb = load_workbook("random.xlsx")
for x in range(1,11):
    for y in range(1,11):
        print(ws.cell(row=y,column=x).value,end=" ")
    print()
lb.close()