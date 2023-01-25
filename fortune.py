from selenium import webdriver
from bs4 import BeautifulSoup
# from fortune2 import Fortune
from openpyxl import Workbook,load_workbook
import requests
import os
import re

inputvalue = '오늘의 운세'


# brow = webdriver.Chrome()
# brow.get(f'https://fortune.nate.com/contents/freeunse/freeunseframe.nate?freeUnseId=today03')
req = requests.get('https://fortune.nate.com/contents/freeunse/freeunseframe.nate?freeUnseId=today03')
req = requests.get('https://fortune.nate.com/contents/freeunse/dayjiji.nate?jijiPage=0')
req.encoding="EUC-KR"

html = req.text
# html = str(html).strip()
html = BeautifulSoup(html,'html.parser')
fo = html.find_all('div',{'class','wrap f_clear'})
# print(fo)
fortune2 = []

if os.path.isfile('fortune.xlsx'):
    wb = load_workbook('fortune.xlsx')
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(['호랑이'])
    ws.append(['year', 'text'])


for idx,tag in enumerate(fo):
    img = tag.find('img')
    #let = tag.select_one(".year_boxrhemdgkg")
    let = tag.find('div',{'class','conFortune'})
    # print(let)
    con_txt = let.select('#con_txt')
    con_txt = con_txt[0].text
    # ws.append(str(con_txt))
    divs = let.select('.today_year')
    for div in divs:
        year = div.select('.year')[0].text
        txt = div.select('div')[0].text
        ws.append([year, txt])


    wb.save('fortune.xlsx')
    wb.close()